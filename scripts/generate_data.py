#!/usr/bin/env python3
"""Generate src/data.ts verbatim from the source workbook and companion document.

Run:  python3 scripts/generate_data.py
Sources (kept in scripts/ so the build is reproducible):
  - scripts/clause_bank.xlsx
  - scripts/companion.docx
"""
import json
import os
import re

import openpyxl
import docx

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
XLSX = os.path.join(HERE, "clause_bank.xlsx")
DOCX = os.path.join(HERE, "companion.docx")
OUT = os.path.join(ROOT, "src", "data.ts")


def clean(v):
    if v is None:
        return ""
    return str(v).replace("\r\n", "\n").replace("\r", "\n").strip()


def paragraphs(text):
    """Split a cell into display paragraphs on blank lines / single newlines."""
    text = clean(text)
    if not text:
        return []
    parts = re.split(r"\n\s*\n", text)
    if len(parts) == 1:
        parts = text.split("\n")
    return [p.strip() for p in parts if p.strip()]


# ---------------------------------------------------------------------------
# Clause Bank workbook
# ---------------------------------------------------------------------------
wb = openpyxl.load_workbook(XLSX, data_only=True)

# --- Clause Bank sheet ---
ws = wb["Clause Bank"]
rows = list(ws.iter_rows(values_only=True))

PRIORITY_KEYS = [
    "preservation",
    "serviceAvailability",
    "patronPrivacy",
    "accessibility",
    "resourceSharing",
]

clauses = []
current_theme = None
theme_order = []
for row in rows:
    cells = [clean(c) for c in row]
    while len(cells) < 15:
        cells.append("")
    cid = cells[0]
    if not cid or cid == "Clause ID":
        continue
    # A clause row has a Pro-Vendor column (index 4) populated; theme header rows do not.
    if cells[4] == "" and cells[1] == "" and cells[5] == "":
        # theme group header row (single cell)
        current_theme = cid
        if cid not in theme_order:
            theme_order.append(cid)
        continue
    theme = cells[2] or current_theme or ""
    if theme and theme not in theme_order:
        theme_order.append(theme)
    priorities = {k: cells[10 + i] for i, k in enumerate(PRIORITY_KEYS)}
    clauses.append(
        {
            "id": cid,
            "topic": cells[1],
            "theme": theme,
            "section": cells[3],
            "proVendor": clean(cells[4]),
            "proVendorParas": paragraphs(cells[4]),
            "proLibrary": clean(cells[5]),
            "proLibraryParas": paragraphs(cells[5]),
            "fallback": clean(cells[6]),
            "fallbackParas": paragraphs(cells[6]),
            "why": clean(cells[7]),
            "watchFor": clean(cells[8]),
            "companionRef": cells[9],
            "priorities": priorities,
        }
    )

# --- Read Me First sheet ---
readme_ws = wb["Read Me First"]
readme_lines = [clean(r[0]) for r in readme_ws.iter_rows(values_only=True)]
readme_lines = [l for l in readme_lines if l]

README_HEADINGS = {
    "Intro",
    "Notes on Organization: Rows and Columns",
    "Notes on Organization: Priority",
    "How You Can Use This Resource",
    "Working with the Risk Exposure Scorecard",
    "Color Coding",
    "IMPORTANT!",
    "About",
}

readme_title = readme_lines[0]
readme_subtitle = readme_lines[1]
readme_sections = []
cur = None
for line in readme_lines[2:]:
    if line in README_HEADINGS:
        cur = {"heading": line, "body": []}
        readme_sections.append(cur)
    else:
        if cur is None:
            cur = {"heading": "", "body": []}
            readme_sections.append(cur)
        cur["body"].append(line)

# --- Theme Index sheet ---
ti_ws = wb["Theme Index"]
ti_all_lines = [clean(r[0]) for r in ti_ws.iter_rows(min_col=1, max_col=1, values_only=True)]
ti_all_lines = [l for l in ti_all_lines if l]
theme_index_title = ti_all_lines[0] if ti_all_lines else "Theme Index"
theme_index_intro = ti_all_lines[1] if len(ti_all_lines) > 1 else ""
theme_index = []
ti_current = None
for row in ti_ws.iter_rows(min_row=4, values_only=True):
    cells = [clean(c) for c in row]
    while len(cells) < 4:
        cells.append("")
    if not any(cells):
        continue
    if cells[0] and cells[0] != "Theme":
        ti_current = cells[0]
    if cells[1] and cells[1] != "Clause ID":
        theme_index.append(
            {"theme": ti_current, "id": cells[1], "topic": cells[2], "section": cells[3]}
        )

# ---------------------------------------------------------------------------
# Companion document
# ---------------------------------------------------------------------------
doc = docx.Document(DOCX)
para_texts = [clean(p.text) for p in doc.paragraphs]
para_texts = [p for p in para_texts if p]

H1 = {
    "Introduction",
    "Negotiating with Vendors",
    "Grant and Scope",
    "Access and Continuity",
    "Financial Terms",
    "Data and Privacy",
    "Risk Allocation",
    "Term and Exit",
    "Dispute Resolution",
    "Governance and Change Management",
}
H2 = {
    "Negotiation Timing",
    "Working Through an Agreement",
    "Leverage",
    "When the Vendor Won't Move",
    "The Frame of the Grant",
    "Who and Where",
    "What Patrons Can Do",
    "Library-Specific Operations",
    "Perpetual Access",
    "Operational Continuity",
    "The Consolidated Privacy Clause",
    "Security and Patron-Created Content",
    "What the Vendor Is Promising",
    "Indemnification and Limitation of Liability",
    "Other Risk Provisions",
    "Term, Renewal, and Termination",
    "Assignment and Change of Control",
    "Amendment and Incorporation",
    "Notices, Audits, and Integration",
}
# normalize curly apostrophes for matching
def norm(s):
    return s.replace("’", "'").strip()


H2 = {norm(h) for h in H2}

# The Risk Allocation chapter has no explicit H1 in the source; its intro paragraph
# begins with this text. We promote a synthetic "Risk Allocation" heading before it.
RISK_INTRO_PREFIX = "The Risk Allocation clauses distribute financial exposure"

companion_title = para_texts[0]
companion_subtitle = para_texts[1]
companion_updated = para_texts[2]

companion = []  # list of {heading, level, body[]}
cur = None
SKIP_FOOTER_PREFIX = ("© 2026 Library Futures", "Licensed for use under CC-BY-4.0")
for line in para_texts[3:]:
    n = norm(line)
    if line.startswith(SKIP_FOOTER_PREFIX[0]) or line.startswith(SKIP_FOOTER_PREFIX[1]):
        continue
    # synthetic Risk Allocation heading
    if line.startswith(RISK_INTRO_PREFIX):
        cur = {"heading": "Risk Allocation", "level": 1, "body": []}
        companion.append(cur)
    if n in H1:
        cur = {"heading": line, "level": 1, "body": []}
        companion.append(cur)
    elif n in H2:
        cur = {"heading": line, "level": 2, "body": []}
        companion.append(cur)
    else:
        if cur is None:
            cur = {"heading": "Introduction", "level": 1, "body": []}
            companion.append(cur)
        cur["body"].append(line)

# ---------------------------------------------------------------------------
# Emit TypeScript
# ---------------------------------------------------------------------------
PRIORITY_LABELS = {
    "preservation": "Preservation",
    "serviceAvailability": "Service Availability",
    "patronPrivacy": "Patron Privacy",
    "accessibility": "Accessibility",
    "resourceSharing": "Resource Sharing",
}

payload = {
    "themeOrder": theme_order,
    "clauses": clauses,
    "readme": {
        "title": readme_title,
        "subtitle": readme_subtitle,
        "sections": readme_sections,
    },
    "themeIndex": theme_index,
    "companion": {
        "title": companion_title,
        "subtitle": companion_subtitle,
        "updated": companion_updated,
        "sections": companion,
    },
    "priorityLabels": PRIORITY_LABELS,
}


def ts_dump(obj):
    return json.dumps(obj, ensure_ascii=False, indent=2)


header = """// AUTO-GENERATED by scripts/generate_data.py — do not edit by hand.
// All clause and companion text is taken verbatim from the source documents:
//   Digital Content Licensing Clause Bank for Libraries (workbook)
//   Digital Licensing Companion Document
// (c) 2026 Library Futures / The Engelberg Center on Innovation Law & Policy. CC-BY-4.0.

export type Priority = "H" | "M" | "";

export interface ClausePriorities {
  preservation: Priority;
  serviceAvailability: Priority;
  patronPrivacy: Priority;
  accessibility: Priority;
  resourceSharing: Priority;
}

export interface Clause {
  id: string;
  topic: string;
  theme: string;
  section: string;
  proVendor: string;
  proVendorParas: string[];
  proLibrary: string;
  proLibraryParas: string[];
  fallback: string;
  fallbackParas: string[];
  why: string;
  watchFor: string;
  companionRef: string;
  priorities: ClausePriorities;
}

export interface ReadmeSection {
  heading: string;
  body: string[];
}

export interface ThemeIndexEntry {
  theme: string;
  id: string;
  topic: string;
  section: string;
}

export interface CompanionSection {
  heading: string;
  level: number;
  body: string[];
}

export const PRIORITY_LABELS: Record<keyof ClausePriorities, string> =
"""

with open(OUT, "w", encoding="utf-8") as f:
    f.write(header)
    f.write(ts_dump(PRIORITY_LABELS) + " as const;\n\n")
    f.write("export const PRIORITY_KEYS: (keyof ClausePriorities)[] = " +
            ts_dump(PRIORITY_KEYS) + ";\n\n")
    f.write("export const THEME_ORDER: string[] = " + ts_dump(theme_order) + ";\n\n")
    f.write("export const CLAUSES: Clause[] = " + ts_dump(clauses) + ";\n\n")
    f.write("export const README_TITLE = " + json.dumps(readme_title, ensure_ascii=False) + ";\n")
    f.write("export const README_SUBTITLE = " + json.dumps(readme_subtitle, ensure_ascii=False) + ";\n")
    f.write("export const README_SECTIONS: ReadmeSection[] = " + ts_dump(readme_sections) + ";\n\n")
    f.write("export const THEME_INDEX_TITLE = " + json.dumps(theme_index_title, ensure_ascii=False) + ";\n")
    f.write("export const THEME_INDEX_INTRO = " + json.dumps(theme_index_intro, ensure_ascii=False) + ";\n")
    f.write("export const THEME_INDEX: ThemeIndexEntry[] = " + ts_dump(theme_index) + ";\n\n")
    f.write("export const COMPANION_TITLE = " + json.dumps(companion_title, ensure_ascii=False) + ";\n")
    f.write("export const COMPANION_SUBTITLE = " + json.dumps(companion_subtitle, ensure_ascii=False) + ";\n")
    f.write("export const COMPANION_UPDATED = " + json.dumps(companion_updated, ensure_ascii=False) + ";\n")
    f.write("export const COMPANION_SECTIONS: CompanionSection[] = " + ts_dump(companion) + ";\n")

# quick sanity report to stdout
print(f"clauses: {len(clauses)}")
print(f"themes:  {theme_order}")
print(f"readme sections: {[s['heading'] for s in readme_sections]}")
print(f"theme index entries: {len(theme_index)}")
print(f"companion sections: {len(companion)}")
print("companion H1:", [s['heading'] for s in companion if s['level'] == 1])
